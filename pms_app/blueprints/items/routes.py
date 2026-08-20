from __future__ import annotations
import csv
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from flask import (
    Response,
    abort,
    current_app,
    flash,
    redirect,
    render_template,
    request,
    url_for,
)
from flask_login import current_user, login_required
from flask_wtf import FlaskForm
from flask_wtf.file import FileField, FileRequired, FileAllowed
from wtforms import SubmitField
from sqlalchemy.exc import SQLAlchemyError
from pms_app.extensions import db
from pms_app.models.contract import Contract
from pms_app.models.item import ContractItem
from pms_app.utils.security import permission_required
from . import bp
from .forms import DeleteForm, ItemForm, ImportExcelForm  # ImportExcelForm را از forms.py ایمپورت کن

def _require_openpyxl():
    try:
        from openpyxl import Workbook, load_workbook
        return Workbook, load_workbook
    except ImportError:
        return None, None

def _to_decimal(v):
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v).strip())
    except (InvalidOperation, ValueError):
        return None

def _to_bool(v) -> bool:
    if v is None or v == "":
        return False
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    return s in ("1", "true", "yes", "on", "y", "بله", "بلی")

def _to_date(v):
    if v is None or v == "":
        return None
    if hasattr(v, "date"):
        try:
            return v.date()
        except Exception:
            pass
    try:
        return datetime.strptime(str(v).strip(), "%Y-%m-%d").date()
    except Exception:
        return None

@bp.route("/contracts/<int:contract_id>/items")
@login_required
@permission_required("items.read")
def items(contract_id: int):
    contract = db.session.get(Contract, contract_id)
    if not contract:
        abort(404)

    q = request.args.get("q", "").strip()
    status = request.args.get("status", "").strip()
    priority = request.args.get("priority", "").strip()

    query = ContractItem.query.filter_by(contract_id=contract_id)

    if q:
        like = f"%{q}%"
        query = query.filter(
            (ContractItem.title.ilike(like)) |
            (ContractItem.wbs_code.ilike(like)) |
            (ContractItem.pms_item_number.ilike(like)) |
            (ContractItem.zone.ilike(like)) |
            (ContractItem.tag_number.ilike(like)) |
            (ContractItem.l1_project_no.ilike(like)) |     # جدید
            (ContractItem.l9_activity_name.ilike(like))     # جدید
        )

    if status:
        query = query.filter(ContractItem.status == status)
    if priority:
        query = query.filter(ContractItem.priority == priority)

    page = request.args.get("page", 1, type=int)
    per_page = current_app.config.get("PER_PAGE", 20)

    pagination = query.order_by(ContractItem.updated_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )

    delete_form = DeleteForm()
    import_form = ImportExcelForm()

    return render_template(
        "items/items.html",
        contract=contract,
        items=pagination.items,
        pagination=pagination,
        q=q,
        status=status,
        priority=priority,
        delete_form=delete_form,
        import_form=import_form,
    )

@bp.route("/contracts/<int:contract_id>/items/template.xlsx")
@login_required
@permission_required("items.read")
def items_template_xlsx(contract_id: int):
    contract = db.session.get(Contract, contract_id)
    if not contract:
        abort(404)

    Workbook, _ = _require_openpyxl()
    if Workbook is None:
        flash("برای دانلود قالب اکسل باید پکیج openpyxl نصب باشد.", "danger")
        return redirect(url_for("items.items", contract_id=contract_id))

    wb = Workbook()
    ws = wb.active
    ws.title = "Items Template"

    # هدرها با نام انگلیسی فیلدهای مدل
    headers = [
        "title", "description", "pms_item_number", "zone", "phase", "area",
        "discipline", "work_package", "wbs_code", "boq_item_code", "activity_id",
        "equipment_number", "unit_of_measure", "tag_number", "currency",
        "cost_center", "original_amount", "adjusted_amount", "weight_factor",
        "planned_quantity", "actual_progress_percentage", "status", "priority",
        "is_common", "workfront", "risk_level", "quality_metrics",
        "acceptance_criteria", "stakeholder_id", "baseline_start_date",
        "baseline_end_date", "actual_start_date", "actual_end_date", "remarks",
        # فیلدهای جدید
        "l1_project_no", "l2_sub_project", "l3_phase", "l4_discipline",
        "l5_area_zone", "l6_site_location", "l7_equipment_tag", "l8_work_package",
        "l9_activity_name", "responsible_owner_id", "is_milestone",
        "approval_status", "approval_date", "revision_number", "estimated_duration",
        "predecessors", "successors", "actual_quantity", "actual_cost",
        "cost_category", "funding_source", "resource_assignment"
    ]

    ws.append(headers)

    # یک ردیف نمونه (اختیاری)
    ws.append([
        "Sample Item Title", "Sample description", "PMS-001", "Zone-A", "Phase-1",
        "Area-01", "Civil", "WP-Excavation", "1.2.3.4", "BOQ-567", "ACT-1001",
        "EQ-EXC01", "m3", "TAG-EXC-01", "USD", "CC-200", "50000.00", "52000.00",
        "1.5", "1200", "0", "open", "medium", "0", "Front-1", "low", "",
        "", "", "2026-03-01", "2026-06-30", "", "", "Sample remarks",
        "PROJ-2026", "SubProj-A", "Execution", "Mechanical", "Zone-North",
        "Site-Main", "TAG-PUMP01", "Package-Piping", "Install Pump",
        "101", "1", "pending", "2026-02-20", "3", "5", "ACT-099,ACT-100",
        "ACT-102", "800", "45000.00", "labor", "Internal Budget", "John Doe, Pump01"
    ])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return Response(
        bio.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="items_template_contract_{contract_id}.xlsx"'},
    )

@bp.route("/contracts/<int:contract_id>/items/import", methods=["POST"])
@login_required
@permission_required("items.write")
def items_import(contract_id: int):
    contract = db.session.get(Contract, contract_id)
    if not contract:
        abort(404)

    form = ImportExcelForm()
    if not form.validate_on_submit():
        flash("فرم نامعتبر یا فایل انتخاب نشده.", "danger")
        return redirect(url_for("items.items", contract_id=contract_id))

    f = form.file.data
    if not f:
        flash("فایل انتخاب نشده است.", "danger")
        return redirect(url_for("items.items", contract_id=contract_id))

    _, load_workbook = _require_openpyxl()
    if load_workbook is None:
        flash("پکیج openpyxl نصب نیست.", "danger")
        return redirect(url_for("items.items", contract_id=contract_id))

    try:
        wb = load_workbook(f, data_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(h).strip().lower() if h is not None else "" for h in header_row]
        idx = {h: i for i, h in enumerate(headers) if h}

        def get(row, key, default=None):
            i = idx.get(key.lower())
            if i is None:
                return default
            val = row[i] if i < len(row) else default
            return val.strip() if isinstance(val, str) else val

        created = 0
        skipped = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue

            title = get(row, "title", "")
            if not title:
                skipped += 1
                continue

            item = ContractItem(
                contract_id=contract_id,
                created_by_id=current_user.id,
                updated_by_id=current_user.id,
            )
            item.title = title
            item.description = get(row, "description")
            item.pms_item_number = get(row, "pms_item_number")
            item.zone = get(row, "zone")
            item.phase = get(row, "phase")
            item.area = get(row, "area")
            item.discipline = get(row, "discipline")
            item.work_package = get(row, "work_package")
            item.wbs_code = get(row, "wbs_code")
            item.boq_item_code = get(row, "boq_item_code")
            item.activity_id = get(row, "activity_id")
            item.equipment_number = get(row, "equipment_number")
            item.unit_of_measure = get(row, "unit_of_measure")
            item.tag_number = get(row, "tag_number")
            item.currency = get(row, "currency")
            item.cost_center = get(row, "cost_center")
            item.original_amount = _to_decimal(get(row, "original_amount"))
            item.adjusted_amount = _to_decimal(get(row, "adjusted_amount"))
            item.weight_factor = _to_decimal(get(row, "weight_factor"))
            item.planned_quantity = _to_decimal(get(row, "planned_quantity"))
            item.actual_progress_percentage = _to_decimal(get(row, "actual_progress_percentage"))
            item.status = (get(row, "status") or "open").strip().lower()
            item.priority = (get(row, "priority") or "medium").strip().lower()
            item.is_common = _to_bool(get(row, "is_common"))
            item.workfront = get(row, "workfront")
            item.risk_level = get(row, "risk_level")
            item.quality_metrics = get(row, "quality_metrics")
            item.acceptance_criteria = get(row, "acceptance_criteria")
            item.stakeholder_id = get(row, "stakeholder_id")
            item.baseline_start_date = _to_date(get(row, "baseline_start_date"))
            item.baseline_end_date = _to_date(get(row, "baseline_end_date"))
            item.actual_start_date = _to_date(get(row, "actual_start_date"))
            item.actual_end_date = _to_date(get(row, "actual_end_date"))
            item.remarks = get(row, "remarks")

            # فیلدهای جدید
            item.l1_project_no = get(row, "l1_project_no")
            item.l2_sub_project = get(row, "l2_sub_project")
            item.l3_phase = get(row, "l3_phase")
            item.l4_discipline = get(row, "l4_discipline")
            item.l5_area_zone = get(row, "l5_area_zone")
            item.l6_site_location = get(row, "l6_site_location")
            item.l7_equipment_tag = get(row, "l7_equipment_tag")
            item.l8_work_package = get(row, "l8_work_package")
            item.l9_activity_name = get(row, "l9_activity_name")
            item.responsible_owner_id = get(row, "responsible_owner_id")
            item.is_milestone = _to_bool(get(row, "is_milestone"))
            item.approval_status = get(row, "approval_status")
            item.approval_date = _to_date(get(row, "approval_date"))
            item.revision_number = get(row, "revision_number")
            item.estimated_duration = _to_decimal(get(row, "estimated_duration"))
            item.predecessors = get(row, "predecessors")
            item.successors = get(row, "successors")
            item.actual_quantity = _to_decimal(get(row, "actual_quantity"))
            item.actual_cost = _to_decimal(get(row, "actual_cost"))
            item.cost_category = get(row, "cost_category")
            item.funding_source = get(row, "funding_source")
            item.resource_assignment = get(row, "resource_assignment")

            db.session.add(item)
            created += 1

        db.session.commit()
        flash(f"آیتم‌ها وارد شدند. ثبت‌شده: {created} | رد شده: {skipped}", "success")

    except SQLAlchemyError as e:
        db.session.rollback()
        current_app.logger.exception("Database error during import")
        flash("خطا در ذخیره در دیتابیس.", "danger")
    except Exception as e:
        current_app.logger.exception("Excel processing error")
        flash("فایل اکسل نامعتبر یا قالب ناسازگار است.", "danger")

    return redirect(url_for("items.items", contract_id=contract_id))

@bp.route("/contracts/<int:contract_id>/items/new", methods=["GET", "POST"])
@login_required
@permission_required("items.write")
def item_new(contract_id: int):
    contract = db.session.get(Contract, contract_id)
    if not contract:
        abort(404)

    form = ItemForm()
    if form.validate_on_submit():
        item = ContractItem(
            contract_id=contract_id,
            created_by_id=current_user.id,
            updated_by_id=current_user.id,
        )
        form.populate_obj(item)
        db.session.add(item)
        try:
            db.session.commit()
            flash("آیتم جدید ایجاد شد.", "success")
            return redirect(url_for("items.items", contract_id=contract_id))
        except SQLAlchemyError:
            db.session.rollback()
            flash("خطا در ذخیره آیتم.", "danger")

    return render_template(
        "items/item_form.html",
        contract=contract,
        form=form,
        title="New Item"
    )

@bp.route("/items/<int:item_id>/edit", methods=["GET", "POST"])
@login_required
@permission_required("items.write")
def item_edit(item_id: int):
    item = db.session.get(ContractItem, item_id)
    if not item:
        abort(404)

    contract = db.session.get(Contract, item.contract_id)
    if not contract:
        abort(404)

    form = ItemForm(obj=item)
    if form.validate_on_submit():
        form.populate_obj(item)
        item.updated_by_id = current_user.id
        try:
            db.session.commit()
            flash("آیتم بروزرسانی شد.", "success")
            return redirect(url_for("items.items", contract_id=item.contract_id))
        except SQLAlchemyError:
            db.session.rollback()
            flash("خطا در بروزرسانی.", "danger")

    return render_template(
        "items/item_form.html",
        contract=contract,
        form=form,
        title="Edit Item"
    )

@bp.route("/items/<int:item_id>/delete", methods=["POST"])
@login_required
@permission_required("items.write")
def item_delete(item_id: int):
    item = db.session.get(ContractItem, item_id)
    if not item:
        abort(404)

    form = DeleteForm()
    if not form.validate_on_submit():
        flash("درخواست نامعتبر.", "danger")
        return redirect(url_for("items.items", contract_id=item.contract_id))

    contract_id = item.contract_id
    try:
        db.session.delete(item)
        db.session.commit()
        flash("آیتم حذف شد.", "info")
    except SQLAlchemyError:
        db.session.rollback()
        flash("خطا در حذف آیتم.", "danger")

    return redirect(url_for("items.items", contract_id=contract_id))

@bp.route("/contracts/<int:contract_id>/items/export.csv")
@login_required
@permission_required("items.read")
def items_export(contract_id: int):
    contract = db.session.get(Contract, contract_id)
    if not contract:
        abort(404)

    items_list = ContractItem.query.filter_by(contract_id=contract_id).order_by(ContractItem.id).all()

    si = StringIO()
    writer = csv.writer(si, quoting=csv.QUOTE_MINIMAL)

    headers = [
        "ID", "Title", "PMS Item Number", "WBS Code", "Work Package", "Zone", "Phase", "Area",
        "Discipline", "BOQ Item Code", "Activity ID", "Equipment Number", "Tag Number", "Unit of Measure",
        "Planned Quantity", "Actual Quantity", "Actual Progress %", "Weight Factor", "Original Amount",
        "Adjusted Amount", "Currency", "Cost Center", "Status", "Priority", "Risk Level",
        "Baseline Start", "Baseline End", "Actual Start", "Actual End", "Estimated Duration",
        "Cost Category", "Actual Cost", "Responsible Owner ID", "Is Milestone", "Is Common", "Remarks"
    ]
    writer.writerow(headers)

    for i in items_list:
        writer.writerow([
            i.id,
            i.title,
            i.pms_item_number,
            i.wbs_code,
            i.work_package,
            i.zone,
            i.phase,
            i.area,
            i.discipline,
            i.boq_item_code,
            i.activity_id,
            i.equipment_number,
            i.tag_number,
            i.unit_of_measure,
            i.planned_quantity,
            i.actual_quantity,
            i.actual_progress_percentage,
            i.weight_factor,
            i.original_amount,
            i.adjusted_amount,
            i.currency,
            i.cost_center,
            i.status,
            i.priority,
            i.risk_level,
            i.baseline_start_date,
            i.baseline_end_date,
            i.actual_start_date,
            i.actual_end_date,
            i.estimated_duration,
            i.cost_category,
            i.actual_cost,
            i.responsible_owner_id,
            "Yes" if i.is_milestone else "No",
            "Yes" if i.is_common else "No",
            i.remarks
        ])

    output = si.getvalue()
    si.close()

    return Response(
        output,
        mimetype="text/csv",
        headers={"Content-Disposition": f'attachment; filename="items_contract_{contract_id}.csv"'}
    )